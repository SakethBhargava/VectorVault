import os
import magic
import docx
from pdfminer.high_level import extract_text
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from PIL import Image
import pytesseract
from typing import Optional, Tuple
import tempfile
from config import settings

def preprocess_image_for_ocr(image_path: str) -> Image:
    """Enhance image for better OCR results"""
    img = Image.open(image_path)
    # Convert to grayscale
    img = img.convert('L')
    # Increase contrast
    img = img.point(lambda x: 0 if x < 128 else 255, '1')
    return img

def extract_text_from_file(file_path: str) -> Tuple[Optional[str], Optional[str]]:
    """Extract text from any supported file type"""
    try:
        mime = magic.Magic(mime=True)
        file_type = mime.from_file(file_path)
        
        # Text Files
        if file_type == 'text/plain':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read(), None
                
        # PDFs
        elif file_type == 'application/pdf':
            text = extract_text(file_path)
            return text if text.strip() else None, None
            
        # Word Documents
        elif file_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
            doc = docx.Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
            return text if text.strip() else None, None
            
        # Excel
        elif file_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            wb = load_workbook(file_path, data_only=True)
            text = []
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    text.append(' '.join(str(cell) for cell in row if cell))
            return '\n'.join(text) if text else None, None
            
        # HTML
        elif file_type == 'text/html':
            with open(file_path, 'r', encoding='utf-8') as f:
                soup = BeautifulSoup(f.read(), 'html.parser')
                return soup.get_text(separator='\n', strip=True), None
                
        # Images (OCR)
        elif file_type.startswith('image/'):
            img = preprocess_image_for_ocr(file_path)
            text = pytesseract.image_to_string(img)
            return text if text.strip() else None, None
            
        # Unsupported type
        else:
            return None, f"Unsupported file type: {file_type}"
            
    except Exception as e:
        return None, f"Error processing file: {str(e)}"
    finally:
        if os.path.exists(file_path):
            os.unlink(file_path)